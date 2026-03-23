import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


SUPPORTED_EXCEL_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")
INVALID_FILENAME_CHARS = r'[<>:"/\\|?*\x00-\x1f]'
OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_SIGNATURE = b"PK\x03\x04"
RESERVED_NAMES = {
	"CON",
	"PRN",
	"AUX",
	"NUL",
	*(f"COM{i}" for i in range(1, 10)),
	*(f"LPT{i}" for i in range(1, 10)),
}


def find_excel_in_directory(directory: Path) -> Path:
	files = [
		p
		for p in directory.iterdir()
		if p.is_file() and p.suffix.lower() in SUPPORTED_EXCEL_SUFFIXES
	]
	if not files:
		raise FileNotFoundError(f"未在目录中找到 Excel 文件: {directory}")
	return sorted(files)[0]


def sanitize_filename(name: str) -> str:
	sanitized = re.sub(INVALID_FILENAME_CHARS, "_", name.strip())
	sanitized = sanitized.rstrip(" .")

	if not sanitized:
		sanitized = "output"

	if sanitized.upper() in RESERVED_NAMES:
		sanitized = f"{sanitized}_file"

	return sanitized


def detect_excel_format(excel_path: Path) -> str:
	with excel_path.open("rb") as f:
		header = f.read(8)

	if header.startswith(ZIP_SIGNATURE):
		return "openxml"
	if header == OLE_SIGNATURE:
		return "ole"

	if excel_path.suffix.lower() == ".xls":
		return "ole"
	return "openxml"


def build_output_name(raw_name: object) -> str:
	if raw_name is None or str(raw_name).strip() == "":
		raise ValueError("A2 单元格为空，无法作为 JSON 文件名")

	output_name = sanitize_filename(str(raw_name))
	if not output_name.lower().endswith(".json"):
		output_name += ".json"
	return output_name


def extract_data_openxml(excel_path: Path) -> tuple[str, list[dict[str, object]]]:
	workbook = load_workbook(excel_path, data_only=True)
	sheet = workbook.active

	raw_name = sheet["A2"].value
	output_name = build_output_name(raw_name)

	headers: list[str] = []
	for col in range(1, 20):  # A-S
		value = sheet.cell(row=3, column=col).value
		header = "" if value is None else str(value).strip()
		headers.append(header)

	rows: list[dict[str, object]] = []
	row_index = 4

	while True:
		first_col_value = sheet.cell(row=row_index, column=1).value
		if first_col_value is None or str(first_col_value).strip() == "":
			break

		row_data: dict[str, object] = {}
		for col, header in enumerate(headers, start=1):
			key = header if header else f"column_{col}"
			row_data[key] = sheet.cell(row=row_index, column=col).value
		rows.append(row_data)
		row_index += 1

	workbook.close()
	return output_name, rows


def extract_data_ole(excel_path: Path) -> tuple[str, list[dict[str, object]]]:
	try:
		import xlrd
	except ImportError as error:
		raise RuntimeError("读取旧版 Excel 需要安装 xlrd，请执行: pip install xlrd") from error

	workbook = xlrd.open_workbook(filename=str(excel_path))
	sheet = workbook.sheet_by_index(0)

	raw_name = sheet.cell_value(1, 0) if sheet.nrows > 1 else ""
	output_name = build_output_name(raw_name)

	headers: list[str] = []
	for col in range(19):  # A-S
		value = sheet.cell_value(2, col) if sheet.nrows > 2 and col < sheet.ncols else ""
		header = "" if value is None else str(value).strip()
		headers.append(header)

	rows: list[dict[str, object]] = []
	row_index = 3  # 第 4 行，xlrd 从 0 开始

	while row_index < sheet.nrows:
		first_col_value = sheet.cell_value(row_index, 0) if sheet.ncols > 0 else ""
		if first_col_value is None or str(first_col_value).strip() == "":
			break

		row_data: dict[str, object] = {}
		for col, header in enumerate(headers):
			key = header if header else f"column_{col + 1}"
			value = sheet.cell_value(row_index, col) if col < sheet.ncols else ""
			row_data[key] = value
		rows.append(row_data)
		row_index += 1

	workbook.release_resources()
	return output_name, rows


def extract_data(excel_path: Path) -> tuple[str, list[dict[str, object]]]:
	excel_format = detect_excel_format(excel_path)
	if excel_format == "ole":
		return extract_data_ole(excel_path)
	return extract_data_openxml(excel_path)


def main() -> int:
	parser = argparse.ArgumentParser(description="提取 Excel 课程信息并输出为 JSON")
	parser.add_argument(
		"excel",
		nargs="?",
		help="可选：Excel 文件路径。不传时默认读取脚本同目录下第一个 Excel 文件。",
	)
	args = parser.parse_args()

	if args.excel:
		excel_path = Path(args.excel).expanduser().resolve()
		if not excel_path.exists():
			print(f"Excel 文件不存在: {excel_path}", file=sys.stderr)
			return 1
	else:
		script_dir = Path(__file__).resolve().parent
		try:
			excel_path = find_excel_in_directory(script_dir)
		except FileNotFoundError as error:
			print(str(error), file=sys.stderr)
			return 1

	try:
		output_name, data = extract_data(excel_path)
	except Exception as error:
		print(f"处理失败: {error}", file=sys.stderr)
		return 1

	output_path = excel_path.with_name(output_name)
	with output_path.open("w", encoding="utf-8") as f:
		json.dump(data, f, ensure_ascii=False, indent=2)

	print(f"已生成 JSON: {output_path}")
	print(f"数据条数: {len(data)}")
	return 0


if __name__ == "__main__":
	raise SystemExit(main())
